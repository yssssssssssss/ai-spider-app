import json
import zipfile
from io import BytesIO
from pathlib import Path
from uuid import UUID

from fastapi import HTTPException
from openpyxl import Workbook
from sqlalchemy.orm import Session

from app import crud
from app.config import settings


def _safe_project_path(path: str | None) -> Path | None:
    if not path:
        return None
    root = Path(settings.PROJECT_ROOT).resolve()
    candidate = Path(path)
    if not candidate.is_absolute():
        candidate = root / candidate
    resolved = candidate.resolve()
    try:
        resolved.relative_to(root)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Unsafe export path") from exc
    return resolved


def _analysis_dict(analysis):
    if not analysis:
        return None
    return {
        "id": str(analysis.id),
        "status": analysis.status,
        "embedding_status": analysis.embedding_status,
        "embedding_error": analysis.embedding_error,
        "design_analysis": analysis.design_analysis,
        "ops_analysis": analysis.ops_analysis,
        "analyzed_at": analysis.analyzed_at.isoformat() if analysis.analyzed_at else None,
    }


def _image_dict(image):
    path = _safe_project_path(image.file_path)
    return {
        "id": str(image.id),
        "task_id": str(image.task_id) if image.task_id else None,
        "task_run_id": str(image.task_run_id) if image.task_run_id else None,
        "device_id": str(image.device_id) if image.device_id else None,
        "file_path": image.file_path,
        "file_exists": bool(path and path.exists()),
        "oss_url": image.oss_url,
        "oss_key": image.oss_key,
        "source_app": image.source_app,
        "scenario": image.scenario,
        "captured_at": image.captured_at.isoformat() if image.captured_at else None,
        "created_at": image.created_at.isoformat() if image.created_at else None,
        "analysis": _analysis_dict(image.analysis),
    }


def task_export_payload(db: Session, task_id: UUID, user_id: UUID | None = None) -> dict:
    task = crud.get_task_for_user(db, task_id, user_id) if user_id else crud.get_task(db, task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    runs = crud.list_task_runs(db, task_id)
    images = crud.list_images(db, task_id=task_id, limit=10000, user_id=user_id)
    return {
        "task": {
            "id": str(task.id),
            "request_id": str(task.request_id) if task.request_id else None,
            "name": task.name,
            "keyword": task.keyword,
            "target_app": task.target_app,
            "target_scenario": task.target_scenario,
            "mode": task.mode,
            "generated_instruction": task.generated_instruction,
            "target_goals_json": task.target_goals_json,
            "status": task.status,
            "created_by": str(task.created_by) if task.created_by else None,
            "approved_by": str(task.approved_by) if task.approved_by else None,
            "run_by": str(task.run_by) if task.run_by else None,
            "approved_at": task.approved_at.isoformat() if task.approved_at else None,
            "completed_at": task.completed_at.isoformat() if task.completed_at else None,
        },
        "runs": [
            {
                "id": str(run.id),
                "attempt_no": run.attempt_no,
                "status": run.status,
                "started_at": run.started_at.isoformat() if run.started_at else None,
                "completed_at": run.completed_at.isoformat() if run.completed_at else None,
                "exit_code": run.exit_code,
                "failure_reason": run.failure_reason,
                "goal_validation_json": run.goal_validation_json,
                "log_path": run.log_path,
                "output_dir": run.output_dir,
                "device_id": str(run.device_id) if run.device_id else None,
            }
            for run in runs
        ],
        "images": [_image_dict(image) for image in images],
    }


def watch_plan_export_payload(db: Session, plan_id: UUID, user_id: UUID | None = None) -> dict:
    plan = crud.get_watch_plan_for_user(db, plan_id, user_id) if user_id else crud.get_watch_plan(db, plan_id)
    if not plan:
        raise HTTPException(status_code=404, detail="Watch plan not found")
    runs = crud.list_watch_runs(db, plan_id, limit=10000)
    reports = crud.list_watch_period_reports(db, plan_id, limit=10000)
    snapshots = []
    for run in runs:
        for snapshot in crud.list_watch_snapshots(db, run.id):
            snapshots.append({
                "id": str(snapshot.id),
                "watch_run_id": str(snapshot.watch_run_id),
                "image_id": str(snapshot.image_id),
                "is_primary": snapshot.is_primary,
                "page_signature": snapshot.page_signature,
                "image": _image_dict(snapshot.image),
            })
    return {
        "plan": {
            "id": str(plan.id),
            "name": plan.name,
            "target_app": plan.target_app,
            "target_page": plan.target_page,
            "entry_instruction": plan.entry_instruction,
            "focus_question": plan.focus_question,
            "capture_scope": plan.capture_scope,
            "schedule_time": str(plan.schedule_time),
            "status": plan.status,
            "pause_reason": plan.pause_reason,
            "last_run_at": plan.last_run_at.isoformat() if plan.last_run_at else None,
            "created_by": str(plan.created_by) if plan.created_by else None,
            "updated_by": str(plan.updated_by) if plan.updated_by else None,
        },
        "runs": [
            {
                "id": str(run.id),
                "task_id": str(run.task_id) if run.task_id else None,
                "run_date": run.run_date.isoformat(),
                "attempt_count": run.attempt_count,
                "status": run.status,
                "failure_reason": run.failure_reason,
                "screenshot_count": run.screenshot_count,
                "valid_snapshot_count": run.valid_snapshot_count,
                "summary": run.daily_summary.summary if run.daily_summary else None,
                "completed_at": run.completed_at.isoformat() if run.completed_at else None,
            }
            for run in runs
        ],
        "snapshots": snapshots,
        "period_reports": [
            {
                "id": str(report.id),
                "period_days": report.period_days,
                "date_from": report.date_from.isoformat(),
                "date_to": report.date_to.isoformat(),
                "report": report.report,
                "structured_json": report.structured_json,
            }
            for report in reports
        ],
    }


def json_bytes(payload: dict) -> bytes:
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def excel_bytes(payload: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "overview"
    root_key = "task" if "task" in payload else "plan"
    ws.append(["field", "value"])
    for key, value in payload[root_key].items():
        ws.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])

    runs = wb.create_sheet("runs")
    run_rows = payload.get("runs", [])
    _append_dict_rows(runs, run_rows)

    images = wb.create_sheet("images")
    image_rows = payload.get("images") or [item["image"] for item in payload.get("snapshots", [])]
    _append_dict_rows(images, [_flatten_image(row) for row in image_rows])

    analyses = wb.create_sheet("analysis")
    _append_dict_rows(analyses, [_flatten_analysis(row) for row in image_rows if row.get("analysis")])

    failures = wb.create_sheet("failures")
    failure_rows = [row for row in run_rows if row.get("failure_reason") or row.get("status") in ("failed", "timeout")]
    _append_dict_rows(failures, failure_rows)

    reports = payload.get("period_reports")
    if reports is not None:
        sheet = wb.create_sheet("period_reports")
        _append_dict_rows(sheet, reports)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _append_dict_rows(sheet, rows: list[dict]):
    if not rows:
        sheet.append(["empty"])
        return
    keys = sorted({key for row in rows for key in row.keys()})
    sheet.append(keys)
    for row in rows:
        sheet.append([_cell_value(row.get(key)) for key in keys])


def _cell_value(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _flatten_image(row: dict) -> dict:
    data = dict(row)
    data.pop("analysis", None)
    return data


def _flatten_analysis(row: dict) -> dict:
    analysis = dict(row.get("analysis") or {})
    analysis["image_id"] = row.get("id")
    analysis["image_path"] = row.get("file_path")
    return analysis


def task_zip_bytes(db: Session, task_id: UUID, user_id: UUID | None = None) -> bytes:
    payload = task_export_payload(db, task_id, user_id=user_id)
    missing = []
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("metadata.json", json_bytes(payload))
        zf.writestr("task.xlsx", excel_bytes(payload))
        for image in payload["images"]:
            path = _safe_project_path(image.get("file_path"))
            if not path or not path.exists():
                missing.append(image.get("file_path"))
                continue
            arcname = f"images/{image['id']}{path.suffix.lower() or '.png'}"
            zf.write(path, arcname)
        zf.writestr("missing_files.json", json_bytes({"missing_files": missing}))
    return buffer.getvalue()
